from openpyxl import load_workbook

class ExcelReader:

    @staticmethod
    def get_test_data(file_path, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]


        data = []
        headers = []

        for index,row in enumerate(sheet.iter_rows(values_only=True)):
            if index == 0:
                headers = row
            else:
               row_dict = dict(zip(headers,row))
               data.append(row_dict)

        return data